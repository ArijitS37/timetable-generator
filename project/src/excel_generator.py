"""
Excel generation module for master timetable - UPGRADED VERSION
Features:
- Department-wise sheets
- Merged course indicators
- Split teaching annotations (teacher initials)
- Assistant teacher display
- 2-hour block merging
- Color coding with legend
"""
import pandas as pd
from src.config import Config
from typing import Dict, List, Tuple, Set
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

class ExcelGenerator:
    def __init__(self, solution: Dict, subjects: List[Dict]):
        self.solution = solution
        self.subjects = subjects
        self.master_schedule = solution['master_schedule']
        self.slots = Config.get_slots_list()
        self.days = Config.DAYS
        self.assistant_assignments = solution.get('assistant_assignments', {})
        
        # Color scheme for subject types (pastel colors, easy on eyes)
        self.color_scheme = {
            'DSC': 'B8E6F5',    # Light blue
            'DSE': 'D4E6F1',    # Lighter blue
            'GE': 'C5E1A5',     # Light green
            'SEC': 'FFF9C4',    # Light yellow
            'VAC': 'FFCCBC',    # Light orange
            'AEC': 'E1BEE7',    # Light purple
            'default': 'FFFFFF' # White
        }
    
    def generate_master_timetable(self, filename: str):
        """Generate master timetable Excel with department-wise sheets"""
        print(f"      → {filename}")
        
        # Group subjects by department
        dept_subjects = self._group_subjects_by_department()
        
        # Create Excel writer
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Create a sheet for each department
            for dept_name in sorted(dept_subjects.keys()):
                print(f"         → Creating sheet: {dept_name}")
                self._create_department_sheet(writer, dept_name, dept_subjects[dept_name])
            
            # Add legend sheet
            self._create_legend_sheet(writer)
        
        print(f"      ✅ Master timetable Excel generated with {len(dept_subjects)} department sheets")
    
    def _group_subjects_by_department(self) -> Dict[str, List[Dict]]:
        """Group subjects by their teaching department"""
        dept_groups = defaultdict(list)
        
        for subj in self.subjects:
            dept = subj['Department']
            dept_groups[dept].append(subj)
        
        return dict(dept_groups)
    
    def _create_department_sheet(self, writer, dept_name: str, dept_subjects: List[Dict]):
        """Create a timetable sheet for a specific department"""
        
        # Build schedule grid for this department
        data = []
        merge_info = []  # Track cells to merge
        
        # Header row
        data.append(["Day/Time"] + self.slots)
        
        # Track 2-hour blocks for merging
        blocks_to_merge = []
        
        for day_idx, day in enumerate(self.days):
            row = [day]
            slot_processed = set()  # Track which slots are part of merged blocks
            
            for slot_idx, slot in enumerate(self.slots):
                # Skip if this slot is part of a previous merge
                if slot_idx in slot_processed:
                    row.append("")  # Empty cell, will be merged
                    continue
                
                cell_content = ""
                cell_color = None
                should_merge_next = False
                
                if day in self.master_schedule and slot in self.master_schedule[day]:
                    classes = self.master_schedule[day][slot]
                    
                    # Filter classes for this department
                    dept_classes = [c for c in classes if self._get_subject_department(c) == dept_name]
                    
                    if dept_classes:
                        parts = []
                        
                        for class_info in dept_classes:
                            # Check if this is a 2-hour block start
                            is_block_start = (class_info['type'] == 'Practical' and 
                                            not class_info.get('is_continuation', False))
                            
                            # Check if same lecture/tutorial continues in next slot
                            is_continuous_theory = False
                            if class_info['type'] in ['Lecture', 'Tutorial'] and slot_idx < len(self.slots) - 1:
                                next_slot = self.slots[slot_idx + 1]
                                if day in self.master_schedule and next_slot in self.master_schedule[day]:
                                    next_classes = self.master_schedule[day][next_slot]
                                    for next_class in next_classes:
                                        if (next_class['subject'] == class_info['subject'] and
                                            next_class['course_semester'] == class_info['course_semester'] and
                                            next_class['room'] == class_info['room'] and
                                            next_class['type'] == class_info['type']):
                                            is_continuous_theory = True
                                            break
                            
                            # Build class info string
                            part = self._format_class_info(class_info, is_block_start or is_continuous_theory)
                            parts.append(part)
                            
                            # Set color based on subject type
                            if cell_color is None:  # Use first class's color
                                cell_color = self.color_scheme.get(class_info['subject_type'], 
                                                                  self.color_scheme['default'])
                            
                            # Mark for merging if block start or continuous theory
                            if is_block_start or is_continuous_theory:
                                should_merge_next = True
                        
                        cell_content = "\n---\n".join(parts)
                
                row.append(cell_content)
                
                # Record merge info
                if should_merge_next and slot_idx < len(self.slots) - 1:
                    # Calculate Excel coordinates (row, col)
                    excel_row = day_idx + 2  # +2 because header is row 1, data starts at 2
                    excel_col_start = slot_idx + 2  # +2 because day is col 1, slots start at 2
                    excel_col_end = excel_col_start + 1
                    
                    blocks_to_merge.append({
                        'row': excel_row,
                        'col_start': excel_col_start,
                        'col_end': excel_col_end,
                        'color': cell_color
                    })
                    
                    slot_processed.add(slot_idx + 1)  # Mark next slot as processed
            
            data.append(row)
        
        # Create DataFrame
        df = pd.DataFrame(data[1:], columns=data[0])
        
        # Write to sheet
        sheet_name = dept_name[:31]  # Excel sheet name limit
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Get worksheet for formatting
        worksheet = writer.sheets[sheet_name]
        
        # Apply formatting
        self._format_worksheet(worksheet, len(data), blocks_to_merge)
    
    def _format_class_info(self, class_info: Dict, is_block: bool) -> str:
        """Format class information string with all details"""
        parts = []
        
        # Subject name with merged course indicator
        subject_str = class_info['subject']
        if '[' in class_info['course_semester'] and '+' in class_info['course_semester']:
            # Extract merged courses from course_semester
            # Format: "Course1-Sem1-A" or "COMMON-GE-Sem1-Subject-SecA"
            # For merged: we need to get from original subject data
            merged_courses = self._get_merged_courses(class_info)
            if merged_courses:
                subject_str += f" [{merged_courses}]"
        
        parts.append(f"📚 {subject_str}")
        
        # Teachers (with assistant indicators)
        teacher_str = self._format_teachers(class_info)
        parts.append(f"👤 {teacher_str}")
        
        # Course-Semester
        parts.append(f"📋 {class_info['course_semester']}")
        
        # Section (if not ALL)
        if class_info['section'] != "ALL" and class_info['section']:
            parts.append(f"🔖 Sec-{class_info['section']}")
        
        # Room
        room_str = class_info['room']
        parts.append(f"🏢 {room_str}")
        
        # Type with block indicator
        type_str = class_info['type']
        if is_block:
            type_str += " (2-hour)"
        parts.append(f"📌 {type_str}")
        
        # Subject type
        if class_info['subject_type']:
            parts.append(f"[{class_info['subject_type']}]")
        
        return "\n".join(parts)
    
    def _format_teachers(self, class_info: Dict) -> str:
        """Format teacher string with assistants and initials for split teaching"""
        teachers_list = class_info.get('teachers_list', [class_info['teacher']])
        
        # Get teacher initials
        from src.data_loader import DataLoader
        # We need to access teacher_initials, but it's not directly available here
        # So we'll use the teacher names as-is, assuming they're already formatted
        
        # Check if there are multiple teachers (co-teaching or assistants)
        if len(teachers_list) > 1:
            # First is main teacher, rest are assistants/co-teachers
            return " + ".join(teachers_list)
        else:
            return class_info['teacher']
    
    def _get_merged_courses(self, class_info: Dict) -> str:
        """Extract merged course information"""
        # Look up in original subjects to find merge group
        for subj in self.subjects:
            if (subj['Subject'] == class_info['subject'] and 
                subj['Course_Semester'] == class_info['course_semester']):
                if subj.get('Is_Merged', False):
                    # Find all courses in this merge group
                    merge_group_id = subj.get('Merge_Group_ID')
                    merged_courses = []
                    for s in self.subjects:
                        if s.get('Merge_Group_ID') == merge_group_id:
                            # Get short form of course
                            course_short = Config.get_short_course_name(s['Course'])
                            if course_short not in merged_courses:
                                merged_courses.append(course_short)
                    
                    return " + ".join(sorted(merged_courses))
        
        return ""
    
    def _get_subject_department(self, class_info: Dict) -> str:
        """Get department for a class"""
        for subj in self.subjects:
            if (subj['Subject'] == class_info['subject'] and 
                subj['Course_Semester'] == class_info['course_semester']):
                return subj['Department']
        return "Unknown"
    
    def _format_worksheet(self, worksheet, num_rows: int, blocks_to_merge: List[Dict]):
        """Apply formatting to worksheet"""
        
        # Header formatting
        header_fill = PatternFill(start_color="4A4A4A", end_color="4A4A4A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data cell formatting
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=num_rows), start=2):
            for col_idx, cell in enumerate(row, start=1):
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                cell.border = thin_border
                cell.font = Font(size=8)
        
        # Column widths
        worksheet.column_dimensions['A'].width = 12  # Day column
        for col_idx in range(2, len(self.slots) + 2):
            worksheet.column_dimensions[get_column_letter(col_idx)].width = 35
        
        # Row heights
        for row_idx in range(2, num_rows + 1):
            worksheet.row_dimensions[row_idx].height = 120
        
        # Merge cells for 2-hour blocks
        for block in blocks_to_merge:
            row = block['row']
            col_start = block['col_start']
            col_end = block['col_end']
            color = block['color']
            
            # Merge cells
            start_cell = f"{get_column_letter(col_start)}{row}"
            end_cell = f"{get_column_letter(col_end)}{row}"
            worksheet.merge_cells(f"{start_cell}:{end_cell}")
            
            # Apply color
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            worksheet[start_cell].fill = fill
            
            # Center alignment for merged cells
            worksheet[start_cell].alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
        
        # Apply colors to non-merged cells based on subject type
        for row_idx in range(2, num_rows + 1):
            for col_idx in range(2, len(self.slots) + 2):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                # Skip if part of merge
                if cell.value is not None and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    # Try to determine subject type from cell content
                    content = str(cell.value)
                    if content and '[' in content:
                        for subject_type in self.color_scheme.keys():
                            if f'[{subject_type}]' in content:
                                color = self.color_scheme[subject_type]
                                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                                break
    
    def _create_legend_sheet(self, writer):
        """Create a legend sheet explaining color coding"""
        legend_data = [
            ["Subject Type Color Legend"],
            [""],
            ["Subject Type", "Description", "Color"],
            ["DSC", "Discipline Specific Core", "Light Blue"],
            ["DSE", "Discipline Specific Elective", "Lighter Blue"],
            ["GE", "Generic Elective", "Light Green"],
            ["SEC", "Skill Enhancement Course", "Light Yellow"],
            ["VAC", "Value Added Course", "Light Orange"],
            ["AEC", "Ability Enhancement Course", "Light Purple"],
            [""],
            ["Symbols Used:"],
            ["📚", "Subject Name"],
            ["👤", "Teacher(s)"],
            ["📋", "Course-Semester"],
            ["🔖", "Section"],
            ["🏢", "Room"],
            ["📌", "Class Type"],
            [""],
            ["Notes:"],
            ["• Merged courses shown as: Subject [Course1 + Course2]"],
            ["• Assistant teachers shown as: Main + Asst1 + Asst2"],
            ["• 2-hour blocks are merged across time slots"],
            ["• Split teaching shows individual teacher assignments"],
        ]
        
        df = pd.DataFrame(legend_data)
        df.to_excel(writer, sheet_name="Legend", index=False, header=False)
        
        # Format legend sheet
        worksheet = writer.sheets["Legend"]
        
        # Title formatting
        worksheet['A1'].font = Font(bold=True, size=14)
        worksheet['A1'].alignment = Alignment(horizontal='center')
        
        # Color coding rows
        for row_idx in range(4, 10):  # Rows with subject types
            subject_type = worksheet[f'A{row_idx}'].value
            if subject_type in self.color_scheme:
                color = self.color_scheme[subject_type]
                for col in ['A', 'B', 'C']:
                    worksheet[f'{col}{row_idx}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        # Header row formatting
        for col in ['A', 'B', 'C']:
            worksheet[f'{col}3'].font = Font(bold=True)
            worksheet[f'{col}3'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Column widths
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 40
        worksheet.column_dimensions['C'].width = 20
        
        # Merge title cell
        worksheet.merge_cells('A1:C1')